"""
services/excel_service.py
Генерирует .xlsx файлы для экспорта сдач и записей на зачёты.
aiosqlite.Row не имеет .get() — конвертируем в dict через dict(row).
"""

from __future__ import annotations

import asyncio
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

HEADER_BG  = "2C5F8A"
HEADER_FG  = "FFFFFF"
ROW_ALT_BG = "EBF3FB"


def _header_style(cell) -> None:
    cell.font      = Font(bold=True, color=HEADER_FG, size=11)
    cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len   = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def _to_dict(row) -> dict:
    """Безопасно конвертировать aiosqlite.Row или dict в dict."""
    if isinstance(row, dict):
        return row
    try:
        return dict(row)
    except Exception:
        return {}


async def export_submissions(rows: list, output_path: str) -> str:
    loop = asyncio.get_event_loop()
    dicts = [_to_dict(r) for r in rows]
    return await loop.run_in_executor(None, _sync_export_submissions, dicts, output_path)


def _sync_export_submissions(rows: list[dict], output_path: str) -> str:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title       = "Сдачи РТ"
    ws.freeze_panes = "A2"

    headers = [
        "№", "ФИО ученика", "Telegram ID",
        "Группа", "Куратор",
        "Дата сдачи", "Время сдачи",
        "Статус", "PDF файл"
    ]
    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=1, column=col, value=h))

    for i, row in enumerate(rows, 1):
        try:
            dt       = datetime.fromisoformat(str(row.get("submitted_at", "")))
            date_str = dt.strftime("%d.%m.%Y")
            time_str = dt.strftime("%H:%M:%S")
        except Exception:
            date_str = str(row.get("submitted_at", ""))
            time_str = ""

        values = [
            i,
            row.get("student_full_name", ""),
            row.get("student_id", ""),
            row.get("group_title") or "—",
            row.get("curator_name") or "—",
            date_str,
            time_str,
            row.get("status", ""),
            Path(str(row.get("pdf_path", ""))).name,
        ]
        fill = ROW_ALT_BG if i % 2 == 0 else "FFFFFF"
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    _auto_width(ws)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Excel сдач: %s (%d строк)", output_path, len(rows))
    return output_path


async def export_exam_bookings(rows: list, output_path: str) -> str:
    loop  = asyncio.get_event_loop()
    dicts = [_to_dict(r) for r in rows]
    return await loop.run_in_executor(None, _sync_export_exams, dicts, output_path)


def _sync_export_exams(rows: list[dict], output_path: str) -> str:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title        = "Записи на зачёт"
    ws.freeze_panes = "A2"

    headers = [
        "№", "ФИО ученика", "Telegram ID",
        "Группа", "Куратор",
        "Дата зачёта", "Время зачёта",
        "Google Meet", "Статус", "Дата записи"
    ]
    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=1, column=col, value=h))

    for i, row in enumerate(rows, 1):
        try:
            created = datetime.fromisoformat(str(row.get("created_at", ""))).strftime("%d.%m.%Y %H:%M")
        except Exception:
            created = str(row.get("created_at", ""))

        values = [
            i,
            row.get("student_full_name", ""),
            row.get("student_id", ""),
            row.get("group_title") or "—",
            row.get("curator_name") or "—",
            row.get("slot_date", ""),
            row.get("slot_time", ""),
            row.get("google_meet_link") or "—",
            row.get("status", ""),
            created,
        ]
        fill = ROW_ALT_BG if i % 2 == 0 else "FFFFFF"
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    _auto_width(ws)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Excel зачётов: %s (%d строк)", output_path, len(rows))
    return output_path


async def export_practices(rows: list, output_path: str) -> str:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _sync_export_practices, rows, output_path)


def _sync_export_practices(rows: list, output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Практики"
    ws.freeze_panes = "A2"

    headers = ["№", "ФИО", "Telegram ID", "Тема практики",
               "Группа", "Куратор", "Дата", "Время", "Статус", "PDF файл"]

    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell)

    for i, row in enumerate(rows, 1):
        try:
            dt = datetime.fromisoformat(row["submitted_at"])
            date_str = dt.strftime("%d.%m.%Y")
            time_str = dt.strftime("%H:%M:%S")
        except Exception:
            date_str = str(row.get("submitted_at", ""))
            time_str = ""

        def _get(r, key, default="—"):
            try: return r[key] or default
            except Exception: return default

        values = [
            i,
            row["student_full_name"],
            row["student_id"],
            _get(row, "description"),
            _get(row, "group_title"),
            _get(row, "curator_name"),
            date_str,
            time_str,
            _get(row, "status", "submitted"),
            str(Path(row["pdf_path"]).name),
        ]

        fill_color = ROW_ALT_BG if i % 2 == 0 else "FFFFFF"
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    _auto_width(ws)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Excel практик сохранён: %s (%d строк)", output_path, len(rows))
    return output_path
